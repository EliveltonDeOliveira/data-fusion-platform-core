"""Uso/cobertura da terra do MapBiomas, da planilha de Estatísticas da Coleção 11.

Lê a aba `COVERAGE_11` (área em hectares) e produz uma linha por
(geocode, class_id, year). A planilha quebra cada município por bioma — RS tem
Pampa e Mata Atlântica —, então as áreas são somadas entre biomas para dar o
total municipal. Células nulas ou <= 0 são omitidas (ausência de linha = 0 ha).

Geocodes que não são municípios reais (ex.: 4300001 Lagoa dos Patos, 4300002
Lagoa Mirim — códigos que o IBGE atribui a corpos d'água) não estão na malha de
municípios e são pulados, com contagem reportada.
"""

from __future__ import annotations

import re
from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

_YEAR_COL = re.compile(r"y(\d{4})")
COVERAGE_SHEET = "COVERAGE_11"

LandUseRow = tuple[str, int, int, float]  # geocode, class_id, year, area_ha
_Key = tuple[str, int, int]


@dataclass
class LandUseStats:
    rows_read: int = 0
    rows_kept: int = 0  # linhas da planilha do RS, município conhecido, class_id conhecido
    skipped_other_state: int = 0
    cells_summed: int = 0  # células (município, bioma, classe, ano) somadas
    rows_emitted: int = 0  # linhas agregadas geradas
    municipios: set[str] = field(default_factory=set)
    biomes: set[str] = field(default_factory=set)
    years: set[int] = field(default_factory=set)
    unknown_class_ids: set[int] = field(default_factory=set)
    unknown_geocodes: set[str] = field(default_factory=set)

    @property
    def ok(self) -> bool:
        return not self.unknown_class_ids and self.rows_emitted > 0


def stream_land_use(
    xlsx_path: Path,
    *,
    state_name: str,
    legend_class_ids: set[int],
    known_geocodes: set[str],
    stats: LandUseStats,
) -> Iterator[LandUseRow]:
    totals = _read_totals(
        xlsx_path,
        state_name=state_name,
        legend_class_ids=legend_class_ids,
        known_geocodes=known_geocodes,
        stats=stats,
    )
    for (geocode, class_id, year), area in sorted(totals.items()):
        stats.rows_emitted += 1
        yield (geocode, class_id, year, area)


def _read_totals(
    xlsx_path: Path,
    *,
    state_name: str,
    legend_class_ids: set[int],
    known_geocodes: set[str],
    stats: LandUseStats,
) -> dict[_Key, float]:
    totals: dict[_Key, float] = {}
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[COVERAGE_SHEET]
        source = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(source)]
        idx = {name: i for i, name in enumerate(header)}

        year_cols = [
            (int(m.group(1)), idx[name]) for name in header if (m := _YEAR_COL.fullmatch(name))
        ]
        if not year_cols:
            raise ValueError(f"nenhuma coluna de ano (yXXXX) em {COVERAGE_SHEET}")
        ci_state, ci_geo, ci_class = idx["state"], idx["geocode"], idx["class"]
        ci_biome = idx.get("biome")

        for raw in source:
            stats.rows_read += 1
            if str(raw[ci_state]).strip() != state_name:
                stats.skipped_other_state += 1
                continue

            geocode = str(raw[ci_geo]).strip()
            if geocode not in known_geocodes:
                stats.unknown_geocodes.add(geocode)
                continue

            try:
                class_id = int(raw[ci_class])
            except (TypeError, ValueError):
                continue
            if class_id not in legend_class_ids:
                stats.unknown_class_ids.add(class_id)
                continue

            stats.rows_kept += 1
            stats.municipios.add(geocode)
            if ci_biome is not None and raw[ci_biome]:
                stats.biomes.add(str(raw[ci_biome]).strip())

            for year, ci in year_cols:
                value = raw[ci]
                if value is None:
                    continue
                area = float(value)
                if area <= 0.0:
                    continue
                stats.years.add(year)
                stats.cells_summed += 1
                key: _Key = (geocode, class_id, year)
                totals[key] = totals.get(key, 0.0) + area
    finally:
        wb.close()

    return totals
